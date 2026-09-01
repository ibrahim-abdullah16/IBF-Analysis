"""
IBF AUC Component Ablation Comparison
=====================================

Purpose
-------
Re-run the threshold-free component ablation on the current IBF-Analysis
pipeline for all three cyclones:

    1. Composite IBF score
    2. Vulnerability alone
    3. Hazard-only weighted score

The ROC/AUC calculation follows the current validation notebook:

    - forecast and DDM records are merged by District + Upazila
    - the observed target is binary: DDM damage > 0
    - the raw continuous predictor score is used directly
    - no forecast severity threshold is applied
    - no DDM Kappa cut-point is applied
    - ROC is swept over every unique predictor value
    - AUC is calculated by the trapezoidal rule

Sector predictors
-----------------
Housing composite:
    Norm_Impact_House

Agriculture composite:
    Norm_Impact_fAPAR

Vulnerability alone:
    Norm_Vul

Hazard-only:
    weighted normalized wind gust + rainfall + storm surge

The hazard weights follow the event/lead weighting used by the IBF
methodology:

Remal:
    1dlt, 2dlt, 3dlt = 0.35 wind + 0.35 rainfall + 0.30 surge

Midhili:
    1dlt, 2dlt = 0.35 wind + 0.35 rainfall + 0.30 surge
    3dlt       = 0.50 wind + 0.50 rainfall + 0.00 surge

Sitrang:
    1dlt, 2dlt = 0.40 wind + 0.20 rainfall + 0.40 surge
    3dlt       = 0.50 wind + 0.50 rainfall + 0.00 surge

Observed sector target
----------------------
Housing:
    No_Total > 0

Agriculture:
    Total_Loss_Land > 0

The script also checks whether the second DDM damage field in each sector
has the same binary impacted/non-impacted pattern. This is useful because
the repository's main AUC analysis reports both sector damage fields.

Output
------
One Excel workbook:

outputs/Ablation_AUC_Comparison/
    All_Cyclones_AUC_Ablation_Comparison.xlsx

Main sheet columns match the requested comparison table:

Cyclone | Lead | Sector | Composite AUC | Vulnerability alone AUC |
Hazard-only (weighted) AUC | Composite wins?

No CSV files are created.
"""

from pathlib import Path
import warnings

import numpy as np
import pandas as pd
import yaml

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ============================================================
# SETTINGS
# ============================================================

CYCLONE_KEYS = ["remal", "midhili", "sitrang"]
LEADS = ["1dlt", "2dlt", "3dlt"]

# Event/lead hazard weights used for the hazard-only ablation.
# Tuple order: (wind, rainfall, storm surge)
HAZARD_WEIGHTS = {
    "remal": {
        "1dlt": (0.35, 0.35, 0.30),
        "2dlt": (0.35, 0.35, 0.30),
        "3dlt": (0.35, 0.35, 0.30),
    },
    "midhili": {
        "1dlt": (0.35, 0.35, 0.30),
        "2dlt": (0.35, 0.35, 0.30),
        "3dlt": (0.50, 0.50, 0.00),
    },
    "sitrang": {
        "1dlt": (0.40, 0.20, 0.40),
        "2dlt": (0.40, 0.20, 0.40),
        "3dlt": (0.50, 0.50, 0.00),
    },
}

PREDICTOR_COLUMNS = {
    "Housing": {
        "Composite": "Norm_Impact_House",
        "Vulnerability alone": "Norm_Vul",
        "Hazard-only (weighted)": "Weighted_Hazard",
    },
    "Agriculture": {
        "Composite": "Norm_Impact_fAPAR",
        "Vulnerability alone": "Norm_Vul",
        "Hazard-only (weighted)": "Weighted_Hazard",
    },
}

PRIMARY_DAMAGE_COLUMN = {
    "Housing": "No_Total",
    "Agriculture": "Total_Loss_Land",
}

SECONDARY_DAMAGE_COLUMN = {
    "Housing": "Amt_Total",
    "Agriculture": "Total_Loss_Amt",
}


# ============================================================
# EXCEL STYLES
# ============================================================

NAVY = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREY = "F2F2F2"
GREEN = "C6E0B4"
AMBER = "FFE699"
RED = "F4CCCC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="C7C7C7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def solid_fill(color):
    return PatternFill("solid", fgColor=color)


def header_font():
    return Font(name="Arial", size=9, bold=True, color=WHITE)


def body_font(bold=False):
    return Font(name="Arial", size=9, bold=bold, color="000000")


# ============================================================
# PROJECT ROOT
# ============================================================


def find_project_root():
    """Find the IBF-Analysis root from script location or current folder."""

    candidates = []

    try:
        script_path = Path(__file__).resolve()
        candidates.extend([script_path.parent, *script_path.parents])
    except NameError:
        pass

    cwd = Path.cwd().resolve()
    candidates.extend([cwd, *cwd.parents])

    seen = set()

    for folder in candidates:
        folder = folder.resolve()

        if folder in seen:
            continue

        seen.add(folder)

        if (
            (folder / "configs").is_dir()
            and (folder / "data").is_dir()
        ):
            return folder

    raise RuntimeError(
        "Could not locate the IBF-Analysis repository root. "
        "Place this script under IBF-Analysis/scripts/ and run it again."
    )


# ============================================================
# ADMIN CLEANING
# ============================================================


def clean_admin(value):
    """Same District/Upazila whitespace cleaning used in the repository."""

    if pd.isna(value):
        return np.nan

    return " ".join(str(value).strip().split())


# ============================================================
# ROC / AUC
# Same first-principles implementation used in current notebook
# ============================================================


def roc_curve(scores, labels):
    """
    Sweep every unique continuous score as a candidate threshold.

    Parameters
    ----------
    scores : array-like
        Continuous forecast predictor.
    labels : array-like
        Binary observed target, 0 or 1.
    """

    scores = np.asarray(scores, dtype=float)
    labels = np.asarray(labels, dtype=int)

    positives = labels.sum()
    negatives = len(labels) - positives

    if positives == 0 or negatives == 0:
        return None, None

    thresholds = np.sort(np.unique(scores))[::-1]
    thresholds = np.concatenate(([np.inf], thresholds, [-np.inf]))

    tpr_list = []
    fpr_list = []

    for threshold in thresholds:
        predicted_positive = scores >= threshold

        tp = np.sum(predicted_positive & (labels == 1))
        fp = np.sum(predicted_positive & (labels == 0))

        tpr_list.append(tp / positives)
        fpr_list.append(fp / negatives)

    return np.array(fpr_list), np.array(tpr_list)


def auc_score(fpr, tpr):
    """Trapezoidal area under the ROC curve, matching current notebook."""

    if fpr is None:
        return np.nan

    order = np.argsort(fpr)

    x = fpr[order]
    y = tpr[order]

    return float(
        np.sum(
            (x[1:] - x[:-1])
            * (y[1:] + y[:-1])
            / 2.0
        )
    )


def calculate_auc(scores, labels):
    """Calculate AUC after dropping rows with missing predictor score."""

    scores = pd.to_numeric(scores, errors="coerce")
    labels = pd.Series(labels).astype(int)

    valid = scores.notna()

    scores_valid = scores[valid].values
    labels_valid = labels[valid].values

    if len(scores_valid) == 0:
        return np.nan, 0, 0, 0

    fpr, tpr = roc_curve(scores_valid, labels_valid)
    auc = auc_score(fpr, tpr)

    n = len(scores_valid)
    positive = int(labels_valid.sum())
    negative = int(n - positive)

    return auc, n, positive, negative


# ============================================================
# LOAD FORECAST
# ============================================================


def load_forecast(file_path, sheet_name, weights):
    """Load one forecast workbook and construct the hazard-only score."""

    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required = [
        "District",
        "Upazila",
        "Norm_Impact_House",
        "Norm_Impact_fAPAR",
        "Norm_Vul",
        "Norm_Wind Gust",
        "Norm_Rainfall",
        "Norm_Storm Surge",
    ]

    missing = [
        column
        for column in required
        if column not in df.columns
    ]

    if missing:
        raise ValueError(
            f"Missing forecast columns in {file_path}: "
            + ", ".join(missing)
        )

    numeric_columns = [
        "Norm_Impact_House",
        "Norm_Impact_fAPAR",
        "Norm_Vul",
        "Norm_Wind Gust",
        "Norm_Rainfall",
        "Norm_Storm Surge",
    ]

    for column in numeric_columns:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    wind_weight, rainfall_weight, surge_weight = weights

    df["Weighted_Hazard"] = (
        wind_weight * df["Norm_Wind Gust"]
        + rainfall_weight * df["Norm_Rainfall"]
        + surge_weight * df["Norm_Storm Surge"]
    )

    df = (
        df
        .dropna(subset=["District", "Upazila"])
        .drop_duplicates(
            subset=["District", "Upazila"],
            keep="first",
        )
        .reset_index(drop=True)
    )

    for column in ["District", "Upazila"]:
        df[column] = df[column].apply(clean_admin)

    return df


# ============================================================
# LOAD DDM
# Same extraction used in current validation notebook
# ============================================================


def load_ddm(ddm_file, house_sheet, agriculture_sheet):
    # --------------------------------------------------------
    # HOUSE
    # --------------------------------------------------------

    raw_house = pd.read_excel(
        ddm_file,
        sheet_name=house_sheet,
        header=None,
    )

    house = (
        raw_house
        .iloc[2:, [0, 1, 5, 9]]
        .copy()
        .reset_index(drop=True)
    )

    house.columns = [
        "District",
        "Upazila",
        "No_Total",
        "Amt_Total",
    ]

    house["District"] = house["District"].ffill()

    house = house.dropna(
        subset=["District", "Upazila"]
    )

    for column in ["No_Total", "Amt_Total"]:
        house[column] = (
            pd.to_numeric(
                house[column],
                errors="coerce",
            )
            .fillna(0)
        )

    for column in ["District", "Upazila"]:
        house[column] = house[column].apply(clean_admin)

    # --------------------------------------------------------
    # AGRICULTURE
    # --------------------------------------------------------

    raw_agriculture = pd.read_excel(
        ddm_file,
        sheet_name=agriculture_sheet,
        header=None,
    )

    agriculture = (
        raw_agriculture
        .iloc[2:, [0, 1, 6, 7]]
        .copy()
        .reset_index(drop=True)
    )

    agriculture.columns = [
        "District",
        "Upazila",
        "Total_Loss_Land",
        "Total_Loss_Amt",
    ]

    agriculture["District"] = agriculture["District"].ffill()

    agriculture = agriculture.dropna(
        subset=["District", "Upazila"]
    )

    for column in ["Total_Loss_Land", "Total_Loss_Amt"]:
        agriculture[column] = (
            pd.to_numeric(
                agriculture[column],
                errors="coerce",
            )
            .fillna(0)
        )

    for column in ["District", "Upazila"]:
        agriculture[column] = agriculture[column].apply(clean_admin)

    return house, agriculture


# ============================================================
# CONSISTENCY CHECKS
# ============================================================


def check_secondary_binary_target(merged, sector):
    """
    Check whether the second damage measure has the same impacted vs
    non-impacted pattern as the primary sector measure.
    """

    primary = PRIMARY_DAMAGE_COLUMN[sector]
    secondary = SECONDARY_DAMAGE_COLUMN[sector]

    primary_binary = (
        pd.to_numeric(merged[primary], errors="coerce")
        .fillna(0)
        > 0
    )

    secondary_binary = (
        pd.to_numeric(merged[secondary], errors="coerce")
        .fillna(0)
        > 0
    )

    mismatch = int((primary_binary != secondary_binary).sum())

    return mismatch


def verify_vulnerability_static(forecasts, cyclone_name):
    """Check that the vulnerability field is static across lead times."""

    keys = ["District", "Upazila"]

    base = (
        forecasts["1dlt"][keys + ["Norm_Vul"]]
        .rename(columns={"Norm_Vul": "Norm_Vul_1dlt"})
    )

    messages = []

    for lead in ["2dlt", "3dlt"]:
        other = (
            forecasts[lead][keys + ["Norm_Vul"]]
            .rename(columns={"Norm_Vul": f"Norm_Vul_{lead}"})
        )

        check = pd.merge(
            base,
            other,
            on=keys,
            how="inner",
        )

        if len(check) != len(base):
            messages.append(
                f"{lead}: matched {len(check)} of {len(base)} 1dlt records"
            )
            continue

        difference = (
            pd.to_numeric(check["Norm_Vul_1dlt"], errors="coerce")
            - pd.to_numeric(check[f"Norm_Vul_{lead}"], errors="coerce")
        ).abs()

        differing = int((difference.fillna(np.inf) > 1e-12).sum())

        if differing > 0:
            messages.append(
                f"{lead}: {differing} vulnerability values differ"
            )

    if messages:
        print(
            f"WARNING: {cyclone_name} vulnerability consistency check: "
            + "; ".join(messages)
        )
    else:
        print(
            f"{cyclone_name}: vulnerability consistency PASS"
        )


# ============================================================
# RUN ONE CYCLONE
# ============================================================


def run_cyclone(project_root, cyclone_key):
    config_file = (
        project_root
        / "configs"
        / f"{cyclone_key}.yaml"
    )

    if not config_file.is_file():
        raise FileNotFoundError(
            f"Config file not found: {config_file}"
        )

    with open(config_file, "r", encoding="utf-8") as file:
        config = yaml.safe_load(file)

    cyclone_name = config["cyclone"]["name"]

    forecast_sheet = config["sheets"]["forecast"]
    house_sheet = config["sheets"]["house"]
    agriculture_sheet = config["sheets"]["agriculture"]

    ddm_file = project_root / config["data"]["observed"]

    forecast_files = {
        lead: project_root / config["data"]["forecasts"][lead]
        for lead in LEADS
    }

    print()
    print("=" * 78)
    print(f"{cyclone_name} - AUC COMPONENT ABLATION")
    print("=" * 78)

    # Load DDM once for the cyclone.
    ddm_house, ddm_agriculture = load_ddm(
        ddm_file,
        house_sheet,
        agriculture_sheet,
    )

    forecasts = {}

    for lead in LEADS:
        weights = HAZARD_WEIGHTS[cyclone_key][lead]

        forecasts[lead] = load_forecast(
            forecast_files[lead],
            forecast_sheet,
            weights,
        )

        print(
            f"{lead}: hazard weights "
            f"wind={weights[0]:.2f}, "
            f"rain={weights[1]:.2f}, "
            f"surge={weights[2]:.2f}"
        )

    verify_vulnerability_static(
        forecasts,
        cyclone_name,
    )

    main_rows = []
    audit_rows = []

    for lead in LEADS:
        forecast = forecasts[lead]
        weights = HAZARD_WEIGHTS[cyclone_key][lead]

        sector_sources = {
            "Housing": ddm_house,
            "Agriculture": ddm_agriculture,
        }

        for sector in ["Housing", "Agriculture"]:
            merged = pd.merge(
                forecast,
                sector_sources[sector],
                on=["District", "Upazila"],
                how="inner",
            )

            damage_column = PRIMARY_DAMAGE_COLUMN[sector]

            labels = (
                pd.to_numeric(
                    merged[damage_column],
                    errors="coerce",
                )
                .fillna(0)
                > 0
            ).astype(int)

            mismatch = check_secondary_binary_target(
                merged,
                sector,
            )

            auc_results = {}
            audit_by_predictor = {}

            for predictor_name, predictor_column in (
                PREDICTOR_COLUMNS[sector].items()
            ):
                auc, n, positive, negative = calculate_auc(
                    merged[predictor_column],
                    labels,
                )

                auc_results[predictor_name] = auc

                audit_by_predictor[predictor_name] = {
                    "n": n,
                    "positive": positive,
                    "negative": negative,
                }

            composite_auc = auc_results["Composite"]
            vulnerability_auc = auc_results["Vulnerability alone"]
            hazard_auc = auc_results["Hazard-only (weighted)"]

            # In the supplied AUC comparison table and reviewer wording,
            # "Composite wins?" means Composite AUC > Hazard-only AUC.
            # Vulnerability-alone AUC is reported alongside it but is not
            # part of this specific win flag.
            if pd.isna(composite_auc) or pd.isna(hazard_auc):
                composite_wins = "NA"
            else:
                composite_wins = (
                    "Yes"
                    if composite_auc > hazard_auc
                    else "No"
                )

            main_rows.append({
                "Cyclone": cyclone_name,
                "Lead": lead,
                "Sector": sector,
                "Composite AUC": composite_auc,
                "Vulnerability alone AUC": vulnerability_auc,
                "Hazard-only (weighted) AUC": hazard_auc,
                "Composite wins?": composite_wins,
            })

            # All predictors are expected to have the same valid n in
            # the current sample data. Record each count for auditing.
            audit_rows.append({
                "Cyclone": cyclone_name,
                "Lead": lead,
                "Sector": sector,
                "Primary DDM target": f"{damage_column} > 0",
                "Secondary DDM field": SECONDARY_DAMAGE_COLUMN[sector],
                "Binary target mismatch count": mismatch,
                "Merged rows": len(merged),
                "Composite n": audit_by_predictor["Composite"]["n"],
                "Vulnerability n": audit_by_predictor["Vulnerability alone"]["n"],
                "Hazard-only n": audit_by_predictor["Hazard-only (weighted)"]["n"],
                "Observed positive": int(labels.sum()),
                "Observed negative": int(len(labels) - labels.sum()),
                "Wind weight": weights[0],
                "Rainfall weight": weights[1],
                "Storm surge weight": weights[2],
                "Composite AUC": composite_auc,
                "Vulnerability alone AUC": vulnerability_auc,
                "Hazard-only (weighted) AUC": hazard_auc,
            })

            print(
                f"  {lead:<4} {sector:<11} "
                f"Composite={composite_auc:.3f}  "
                f"Vulnerability={vulnerability_auc:.3f}  "
                f"Hazard={hazard_auc:.3f}  "
                f"Composite wins={composite_wins}"
            )

            if mismatch > 0:
                print(
                    f"    WARNING: {mismatch} rows differ between "
                    f"the two {sector.lower()} binary DDM damage fields."
                )

    return (
        pd.DataFrame(main_rows),
        pd.DataFrame(audit_rows),
    )


# ============================================================
# SUMMARY
# ============================================================


def build_summary(result):
    rows = []

    predictor_columns = {
        "Composite": "Composite AUC",
        "Vulnerability alone": "Vulnerability alone AUC",
        "Hazard-only (weighted)": "Hazard-only (weighted) AUC",
    }

    # Overall and sector summaries.
    for scope_name, subset in [
        ("All sectors", result),
        ("Housing", result[result["Sector"] == "Housing"]),
        ("Agriculture", result[result["Sector"] == "Agriculture"]),
    ]:
        for predictor, column in predictor_columns.items():
            rows.append({
                "Scope": scope_name,
                "Predictor": predictor,
                "Mean AUC": float(subset[column].mean()),
                "Minimum AUC": float(subset[column].min()),
                "Maximum AUC": float(subset[column].max()),
                "n combinations": int(subset[column].notna().sum()),
            })

    # Winner counts, strict maximum. Ties are labelled Tie.
    winner_counts = {
        "Composite": 0,
        "Vulnerability alone": 0,
        "Hazard-only (weighted)": 0,
        "Tie": 0,
    }

    for _, row in result.iterrows():
        values = {
            "Composite": row["Composite AUC"],
            "Vulnerability alone": row["Vulnerability alone AUC"],
            "Hazard-only (weighted)": row["Hazard-only (weighted) AUC"],
        }

        maximum = max(values.values())
        winners = [
            name
            for name, value in values.items()
            if np.isclose(value, maximum, atol=1e-12, rtol=1e-12)
        ]

        if len(winners) == 1:
            winner_counts[winners[0]] += 1
        else:
            winner_counts["Tie"] += 1

    winner_df = pd.DataFrame([
        {
            "Predictor": predictor,
            "Number of best-AUC combinations": count,
        }
        for predictor, count in winner_counts.items()
    ])

    # Direct Composite vs Hazard-only comparison used by the reviewer.
    comparison_rows = []

    for scope_name, subset in [
        ("All sectors", result),
        ("Housing", result[result["Sector"] == "Housing"]),
        ("Agriculture", result[result["Sector"] == "Agriculture"]),
    ]:
        composite_beats_hazard = int(
            (subset["Composite AUC"] > subset["Hazard-only (weighted) AUC"]).sum()
        )
        hazard_beats_composite = int(
            (subset["Hazard-only (weighted) AUC"] > subset["Composite AUC"]).sum()
        )
        ties = int(
            np.isclose(
                subset["Composite AUC"].values,
                subset["Hazard-only (weighted) AUC"].values,
                atol=1e-12,
                rtol=1e-12,
            ).sum()
        )

        comparison_rows.append({
            "Scope": scope_name,
            "Composite beats hazard-only": composite_beats_hazard,
            "Hazard-only beats composite": hazard_beats_composite,
            "Ties": ties,
            "Total combinations": int(len(subset)),
        })

    comparison_df = pd.DataFrame(comparison_rows)

    return pd.DataFrame(rows), winner_df, comparison_df


# ============================================================
# EXCEL WRITER
# ============================================================


def set_column_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_dataframe_table(
    ws,
    df,
    start_row=1,
    auc_columns=None,
    percent_columns=None,
):
    auc_columns = set(auc_columns or [])
    percent_columns = set(percent_columns or [])

    # Header
    for column_index, column_name in enumerate(df.columns, 1):
        cell = ws.cell(
            row=start_row,
            column=column_index,
            value=column_name,
        )

        cell.font = header_font()
        cell.fill = solid_fill(NAVY)
        cell.alignment = CENTER
        cell.border = BORDER

    # Rows
    for row_index, (_, row) in enumerate(
        df.iterrows(),
        start_row + 1,
    ):
        for column_index, column_name in enumerate(df.columns, 1):
            value = row[column_name]

            if isinstance(value, np.generic):
                value = value.item()

            cell = ws.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.font = body_font()
            cell.border = BORDER
            cell.alignment = (
                LEFT
                if column_index in [1, 3]
                else CENTER
            )

            cell.fill = solid_fill(
                LIGHT_GREY
                if row_index % 2 == 0
                else WHITE
            )

            if column_name in auc_columns and pd.notna(value):
                cell.number_format = "0.000"

            if column_name in percent_columns and pd.notna(value):
                cell.number_format = "0.0%"

    ws.freeze_panes = ws.cell(
        row=start_row + 1,
        column=1,
    ).coordinate


def write_excel(
    output_path,
    result,
    audit,
    summary,
    winner_summary,
    composite_hazard_summary,
):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # --------------------------------------------------------
    # MAIN AUC COMPARISON
    # --------------------------------------------------------

    ws = workbook.create_sheet("AUC_Comparison")

    ws.merge_cells("A1:G1")
    ws["A1"] = "IBF-Analysis - Hazard/Vulnerability AUC Ablation"
    ws["A1"].font = Font(
        name="Arial",
        size=12,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = solid_fill(NAVY)
    ws["A1"].alignment = LEFT

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Threshold-free ROC-AUC comparison of Composite, Vulnerability alone, "
        "and weighted Hazard-only predictors"
    )
    ws["A2"].font = Font(
        name="Arial",
        size=9,
        italic=True,
        color="333333",
    )
    ws["A2"].fill = solid_fill(LIGHT_BLUE)
    ws["A2"].alignment = LEFT

    main_columns = [
        "Cyclone",
        "Lead",
        "Sector",
        "Composite AUC",
        "Vulnerability alone AUC",
        "Hazard-only (weighted) AUC",
        "Composite wins?",
    ]

    main_df = result[main_columns].copy()

    write_dataframe_table(
        ws,
        main_df,
        start_row=4,
        auc_columns=[
            "Composite AUC",
            "Vulnerability alone AUC",
            "Hazard-only (weighted) AUC",
        ],
    )

    # Highlight winner status.
    for row_number in range(5, 5 + len(main_df)):
        win_cell = ws.cell(row=row_number, column=7)

        if str(win_cell.value).lower() == "yes":
            win_cell.fill = solid_fill(GREEN)
            win_cell.font = body_font(bold=True)
        elif str(win_cell.value).lower() == "no":
            win_cell.fill = solid_fill(AMBER)

    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 10,
            "C": 16,
            "D": 16,
            "E": 23,
            "F": 25,
            "G": 18,
        },
    )

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    ws_summary = workbook.create_sheet("Summary")

    ws_summary["A1"] = "Mean AUC summary"
    ws_summary["A1"].font = Font(
        name="Arial",
        size=11,
        bold=True,
        color=WHITE,
    )
    ws_summary["A1"].fill = solid_fill(NAVY)

    write_dataframe_table(
        ws_summary,
        summary,
        start_row=3,
        auc_columns=[
            "Mean AUC",
            "Minimum AUC",
            "Maximum AUC",
        ],
    )

    winner_start = 3 + len(summary) + 3

    ws_summary.cell(
        row=winner_start,
        column=1,
        value="Best-AUC counts across the 18 cyclone/lead/sector combinations",
    ).font = Font(name="Arial", size=10, bold=True)

    write_dataframe_table(
        ws_summary,
        winner_summary,
        start_row=winner_start + 1,
    )

    comparison_start = winner_start + len(winner_summary) + 4

    ws_summary.cell(
        row=comparison_start,
        column=1,
        value="Direct Composite vs Hazard-only AUC comparison",
    ).font = Font(name="Arial", size=10, bold=True)

    write_dataframe_table(
        ws_summary,
        composite_hazard_summary,
        start_row=comparison_start + 1,
    )

    set_column_widths(
        ws_summary,
        {
            "A": 22,
            "B": 28,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 18,
        },
    )

    # --------------------------------------------------------
    # AUDIT
    # --------------------------------------------------------

    ws_audit = workbook.create_sheet("Audit")

    write_dataframe_table(
        ws_audit,
        audit,
        start_row=1,
        auc_columns=[
            "Composite AUC",
            "Vulnerability alone AUC",
            "Hazard-only (weighted) AUC",
        ],
    )

    for column_index in range(1, len(audit.columns) + 1):
        column_letter = get_column_letter(column_index)
        ws_audit.column_dimensions[column_letter].width = 20

    ws_audit.column_dimensions["A"].width = 14
    ws_audit.column_dimensions["B"].width = 10
    ws_audit.column_dimensions["C"].width = 15
    ws_audit.column_dimensions["D"].width = 24
    ws_audit.column_dimensions["E"].width = 23

    # --------------------------------------------------------
    # METHOD
    # --------------------------------------------------------

    ws_method = workbook.create_sheet("Method")

    method_lines = [
        (
            "Purpose",
            "Threshold-free component ablation comparing the current composite IBF score "
            "against vulnerability alone and a weighted hazard-only score."
        ),
        (
            "ROC target",
            "Binary DDM impact: observed damage > 0. Housing uses No_Total; "
            "Agriculture uses Total_Loss_Land."
        ),
        (
            "Composite housing predictor",
            "Norm_Impact_House"
        ),
        (
            "Composite agriculture predictor",
            "Norm_Impact_fAPAR"
        ),
        (
            "Vulnerability predictor",
            "Norm_Vul"
        ),
        (
            "Hazard components",
            "Norm_Wind Gust, Norm_Rainfall, Norm_Storm Surge"
        ),
        (
            "Remal weights",
            "1dlt/2dlt/3dlt: wind 0.35, rainfall 0.35, surge 0.30"
        ),
        (
            "Midhili weights",
            "1dlt/2dlt: 0.35/0.35/0.30; 3dlt: 0.50/0.50/0.00"
        ),
        (
            "Sitrang weights",
            "1dlt/2dlt: 0.40/0.20/0.40; 3dlt: 0.50/0.50/0.00"
        ),
        (
            "AUC method",
            "Every unique continuous predictor score is swept as a threshold. "
            "TPR and FPR are calculated at each threshold and AUC is integrated "
            "with the trapezoidal rule, matching the current validation notebook."
        ),
        (
            "Important",
            "No 0.15/0.30/0.70 forecast severity thresholds are used in AUC. "
            "No Kappa DDM Low/High cut-points are used in AUC."
        ),
    ]

    for row_index, (item, description) in enumerate(method_lines, 1):
        c1 = ws_method.cell(row=row_index, column=1, value=item)
        c2 = ws_method.cell(row=row_index, column=2, value=description)

        c1.font = body_font(bold=True)
        c2.font = body_font()

        c1.fill = solid_fill(LIGHT_BLUE)

        c1.border = BORDER
        c2.border = BORDER

        c1.alignment = LEFT
        c2.alignment = LEFT

    ws_method.column_dimensions["A"].width = 30
    ws_method.column_dimensions["B"].width = 95

    workbook.save(output_path)


# ============================================================
# MAIN
# ============================================================


def main():
    project_root = find_project_root()

    print("=" * 78)
    print("IBF AUC COMPONENT ABLATION")
    print("=" * 78)
    print(f"Repository: {project_root}")
    print("Cyclones: Remal, Midhili, Sitrang")
    print("AUC uses raw continuous scores and DDM damage > 0")

    all_results = []
    all_audits = []

    for cyclone_key in CYCLONE_KEYS:
        result, audit = run_cyclone(
            project_root,
            cyclone_key,
        )

        all_results.append(result)
        all_audits.append(audit)

    result = pd.concat(
        all_results,
        ignore_index=True,
    )

    audit = pd.concat(
        all_audits,
        ignore_index=True,
    )

    cyclone_order = {
        "Remal": 0,
        "Midhili": 1,
        "Sitrang": 2,
    }

    lead_order = {
        "1dlt": 0,
        "2dlt": 1,
        "3dlt": 2,
    }

    sector_order = {
        "Housing": 0,
        "Agriculture": 1,
    }

    result["_cyclone_order"] = result["Cyclone"].map(cyclone_order)
    result["_lead_order"] = result["Lead"].map(lead_order)
    result["_sector_order"] = result["Sector"].map(sector_order)

    result = (
        result
        .sort_values([
            "_cyclone_order",
            "_lead_order",
            "_sector_order",
        ])
        .drop(columns=[
            "_cyclone_order",
            "_lead_order",
            "_sector_order",
        ])
        .reset_index(drop=True)
    )

    audit["_cyclone_order"] = audit["Cyclone"].map(cyclone_order)
    audit["_lead_order"] = audit["Lead"].map(lead_order)
    audit["_sector_order"] = audit["Sector"].map(sector_order)

    audit = (
        audit
        .sort_values([
            "_cyclone_order",
            "_lead_order",
            "_sector_order",
        ])
        .drop(columns=[
            "_cyclone_order",
            "_lead_order",
            "_sector_order",
        ])
        .reset_index(drop=True)
    )

    summary, winner_summary, composite_hazard_summary = build_summary(result)

    output_dir = (
        project_root
        / "outputs"
        / "Ablation_AUC_Comparison"
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        output_dir
        / "All_Cyclones_AUC_Ablation_Comparison.xlsx"
    )

    write_excel(
        output_path,
        result,
        audit,
        summary,
        winner_summary,
        composite_hazard_summary,
    )

    print()
    print("=" * 78)
    print("AUC COMPARISON COMPLETED")
    print("=" * 78)

    display_columns = [
        "Cyclone",
        "Lead",
        "Sector",
        "Composite AUC",
        "Vulnerability alone AUC",
        "Hazard-only (weighted) AUC",
        "Composite wins?",
    ]

    print(
        result[display_columns]
        .to_string(
            index=False,
            formatters={
                "Composite AUC": lambda value: f"{value:.3f}",
                "Vulnerability alone AUC": lambda value: f"{value:.3f}",
                "Hazard-only (weighted) AUC": lambda value: f"{value:.3f}",
            },
        )
    )

    print()
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
