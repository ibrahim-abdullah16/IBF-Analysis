"""
Forecasted Impact Engine
========================

Purpose:
    Generate forecast impact input file by integrating:

    - Wind gust forecast
    - Rainfall forecast
    - Storm surge forecast
    - Vulnerability index
    - fAPAR
    - Building exposure

Matching:
    ADM3 PCode

Output:
    Forecasted_Impact_updated.xlsx


Repository structure:

IBF-Analysis/
|
├── data/
│   └── sample/
│       └── impact forecast data sample/
│           ├── windgust.xlsx
│           ├── rainfall.xlsx
│           ├── stormsurge.xlsx
│           ├── vulnerability.xlsx
│           ├── faapar.xlsx
│           ├── building_count.xlsx
│           └── Forecasted_Impact.xlsx
|
├── scripts/
│   └── forecasted_impact_engine.py
|
└── outputs/
    └── Forecasted_Impact_updated.xlsx


"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook



# ==========================================================
# REPOSITORY PATHS
# ==========================================================

ROOT = Path(__file__).resolve().parents[1]


DATA_DIR = (
    ROOT
    /
    "data"
    /
    "sample"
    /
    "impact forecast data sample"
)


OUTPUT_DIR = ROOT / "outputs"

OUTPUT_DIR.mkdir(
    parents=True,
    exist_ok=True
)



# Input files

WIND_FILE = DATA_DIR / "windgust.xlsx"

RAIN_FILE = DATA_DIR / "rainfall.xlsx"

SURGE_FILE = DATA_DIR / "stormsurge.xlsx"


VULNERABILITY_FILE = DATA_DIR / "vulnerability.xlsx"

FAPAR_FILE = DATA_DIR / "faapar.xlsx"

BUILDING_FILE = DATA_DIR / "building_count.xlsx"


DUMMY_IMPACT_FILE = (
    DATA_DIR /
    "Forecasted_Impact_Dummy.xlsx"
)



FORECASTED_IMPACT_FILE = (
    OUTPUT_DIR /
    "Forecasted_Impact.xlsx"
)



# ==========================================================
# COLUMN SETTINGS
# ==========================================================

# PCode in dummy impact file
IMPACT_PCODE_COL = 2     # Column B


# Output columns

VULNERABILITY_COL = 5    # Column E

WIND_COL = 7             # Column G

RAIN_COL = 9             # Column I

SURGE_COL = 11           # Column K

BUILDING_COL = 13        # Column M

FAPAR_COL = 17           # Column Q



# ==========================================================
# PCODE CLEANING
# ==========================================================

def clean_pcode(value):

    if pd.isna(value):
        return None


    value = str(value).strip()


    # remove comma
    value = value.replace(",", "")


    # remove Excel decimal
    if value.endswith(".0"):
        value = value[:-2]


    # remove BD prefix
    if value.upper().startswith("BD"):
        value = value.upper().replace("BD", "")


    return value



# ==========================================================
# WIND GUST PROCESSING
# ==========================================================

def process_windgust():

    df = pd.read_excel(
        WIND_FILE,
        header=None
    )


    pcodes = [
        clean_pcode(x)
        for x in df.iloc[0,1:]
    ]


    data = df.iloc[1:10,1:]


    result = {}


    for idx,pcode in enumerate(pcodes):

        if pcode is None:
            continue


        values = pd.to_numeric(
            data.iloc[:,idx],
            errors="coerce"
        )


        maximum = values.max()


        if pd.isna(maximum):
            maximum = 0


        # m/s to km/h
        result[pcode] = maximum * 3.6


    return result



# ==========================================================
# RAINFALL PROCESSING
# ==========================================================

def process_rainfall():

    df = pd.read_excel(
        RAIN_FILE,
        header=None
    )


    pcodes = [
        clean_pcode(x)
        for x in df.iloc[0,1:]
    ]


    result = {}


    for idx,pcode in enumerate(pcodes):

        if pcode is None:
            continue


        # accumulated rainfall difference
        # row 9 - row 8
        # meter to mm

        r9 = pd.to_numeric(
            df.iloc[9,idx+1],
            errors="coerce"
        )


        r8 = pd.to_numeric(
            df.iloc[8,idx+1],
            errors="coerce"
        )


        r9 = 0 if pd.isna(r9) else r9

        r8 = 0 if pd.isna(r8) else r8


        result[pcode] = (
            (r9-r8)*1000
        )


    return result



# ==========================================================
# STORM SURGE PROCESSING
# ==========================================================

def process_stormsurge():

    df = pd.read_excel(
        SURGE_FILE
    )


    df.columns = [
        str(c).strip()
        for c in df.columns
    ]


    result = {}


    for _,row in df.iterrows():


        pcode = clean_pcode(
            row["ADM3_PCODE"]
        )


        value = pd.to_numeric(
            row["Storm Surge"],
            errors="coerce"
        )


        if pd.isna(value):
            value = 0


        result[pcode] = value


    return result



# ==========================================================
# GENERIC PCODE-VALUE READER
# Used for:
# Vulnerability
# fAPAR
# Building count
# ==========================================================

def process_generic_file(file):

    df = pd.read_excel(
        file
    )


    df.columns = [
        str(c).strip()
        for c in df.columns
    ]


    pcode_column = df.columns[0]

    value_column = df.columns[1]


    result = {}


    for _,row in df.iterrows():


        pcode = clean_pcode(
            row[pcode_column]
        )


        value = pd.to_numeric(
            row[value_column],
            errors="coerce"
        )


        if pcode is not None:


            if pd.isna(value):
                value = 0


            result[pcode] = value


    return result



# ==========================================================
# UPDATE FORECAST IMPACT FILE
# ==========================================================

def update_forecast_file():

    wind = process_windgust()

    rain = process_rainfall()

    surge = process_stormsurge()

    vulnerability = process_generic_file(
        VULNERABILITY_FILE
    )

    fapar = process_generic_file(
        FAPAR_FILE
    )

    building = process_generic_file(
        BUILDING_FILE
    )



    print("\nInput summary")
    print("-----------------------")

    print("Wind:",len(wind))

    print("Rain:",len(rain))

    print("Storm surge:",len(surge))

    print("Vulnerability:",len(vulnerability))

    print("fAPAR:",len(fapar))

    print("Building:",len(building))



    wb = load_workbook(
        DUMMY_IMPACT_FILE
    )


    ws = wb.active



    matched = {

        "wind":0,

        "rain":0,

        "surge":0,

        "vulnerability":0,

        "fapar":0,

        "building":0

    }



    for row in range(
        2,
        ws.max_row+1
    ):


        pcode = clean_pcode(
            ws.cell(
                row,
                IMPACT_PCODE_COL
            ).value
        )


        if pcode in vulnerability:

            ws.cell(
                row,
                VULNERABILITY_COL
            ).value = vulnerability[pcode]

            matched["vulnerability"] += 1



        if pcode in wind:

            ws.cell(
                row,
                WIND_COL
            ).value = wind[pcode]

            matched["wind"] += 1



        if pcode in rain:

            ws.cell(
                row,
                RAIN_COL
            ).value = rain[pcode]

            matched["rain"] += 1



        if pcode in surge:

            ws.cell(
                row,
                SURGE_COL
            ).value = surge[pcode]

            matched["surge"] += 1



        if pcode in building:

            ws.cell(
                row,
                BUILDING_COL
            ).value = building[pcode]

            matched["building"] += 1



        if pcode in fapar:

            ws.cell(
                row,
                FAPAR_COL
            ).value = fapar[pcode]

            matched["fapar"] += 1



    wb.save(
        FORECASTED_IMPACT_FILE
    )


    print("\nMatching summary")
    print("-----------------------")


    for key,value in matched.items():

        print(
            f"{key}: {value}"
        )


    print(
        "\nSaved:"
    )

    print(
        FORECASTED_IMPACT_FILE
    )



# ==========================================================
# RUN
# ==========================================================

if __name__ == "__main__":

    update_forecast_file()