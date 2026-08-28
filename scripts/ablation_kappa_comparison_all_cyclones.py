"""
IBF Hazard/Vulnerability Ablation Kappa Comparison
===================================================

Purpose
-------
Compare the existing composite IBF forecast against:

1. Vulnerability alone
       Excel column F = Norm_Vul

2. Hazard only, weighted
       0.35 * H + 0.35 * J + 0.30 * L

   where:
       H = Norm_Wind Gust
       J = Norm_Rainfall
       L = Norm_Storm Surge

The DDM severity calibration, quadratic-weighted Kappa calculation,
three-lead common-cut search, forecast threshold derivation, admin-name
matching, and tie-breaking logic are kept consistent with the repository's
current DDM Kappa-calibration workflow.

Important ablation rule
-----------------------
The repository has operational forecast severity classes only for the
composite household and agriculture impact scores. Vulnerability-only and
hazard-only scores do not have independent operational class labels.

Therefore, to avoid introducing a second DDM-fitted forecast calibration,
the ablation predictors are classified with the SAME lead-specific,
sector-specific forecast severity boundaries derived from the operational
composite forecast workbook. Only the DDM Low/High cuts are optimized,
exactly as in the repository.

Output
------
One Excel workbook is created for each cyclone:

outputs/<cyclone>/Ablation_Kappa_Comparison/
    <Cyclone>_Ablation_Kappa_Comparison.xlsx

By default the script processes Remal, Midhili and Sitrang in one run.
No CSV files are created.

The Excel workbook contains one table per damage variable in the same layout
as the supplied "required result" example:

Predictor | Lead | n | DDM cut-points (Low / High) | Kappa
          | Exact % | Adjacent % | Off-Cat. %

Exact, Adjacent and Off-Cat. are mutually exclusive and sum to 100%.
"""

from pathlib import Path
import argparse
import warnings

import numpy as np
import pandas as pd
import yaml

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter, column_index_from_string

warnings.filterwarnings("ignore")


# ============================================================
# CONSTANTS
# ============================================================

LEADS = ["1dlt", "2dlt", "3dlt"]

SAT_IDX = {
    "No Impact": 0,
    "Low": 1,
    "Moderate": 2,
    "High": 3,
}

DDM_IDX = {
    "No Impact": 0,
    "Low": 1,
    "Medium": 2,
    "High": 3,
}

PREDICTORS = [
    ("Composite", None),
    ("Vulnerability alone", "Norm_Vul"),
    ("Hazard-only (weighted)", "Weighted_Hazard"),
]


# ============================================================
# EXCEL STYLES
# ============================================================

NAVY = "1F3864"
LBLUE = "D6E4F0"
WHITE = "FFFFFF"
LGREY = "F2F2F2"
GREEN = "C8E6C9"
AMBER = "FFE0B2"
RED_BG = "FFCDD2"

thin = Side(style="thin", color="CCCCCC")
TB = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)

LEFT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
)


def header_font():
    return Font(
        name="Arial",
        size=9,
        bold=True,
        color=WHITE,
    )


def body_font(bold=False):
    return Font(
        name="Arial",
        size=9,
        bold=bold,
        color="000000",
    )


def fill(hex_color):
    return PatternFill(
        "solid",
        fgColor=hex_color,
    )


def kappa_fill(value):
    if pd.isna(value):
        return WHITE
    if value >= 0.60:
        return GREEN
    if value >= 0.40:
        return AMBER
    return RED_BG


# ============================================================
# HELPERS
# ============================================================

def clean_admin(x):
    """
    Same admin-name cleaning used in the repository DDM workflow.
    """
    if pd.isna(x):
        return np.nan

    return " ".join(
        str(x)
        .strip()
        .split()
    )


def weighted_kappa(a, b, K=4):
    """
    Same quadratic-weighted Kappa implementation as the repository.
    """
    a = np.asarray(a)
    b = np.asarray(b)

    if len(a) == 0:
        return np.nan

    O = np.zeros(
        (K, K),
        dtype=float,
    )

    for x, y in zip(a, b):
        O[int(x), int(y)] += 1

    N = O.sum()

    if N == 0:
        return np.nan

    row_totals = O.sum(axis=1)
    col_totals = O.sum(axis=0)

    W = np.array([
        [
            ((i - j) ** 2) / (K - 1) ** 2
            for j in range(K)
        ]
        for i in range(K)
    ])

    observed = (W * O).sum()

    expected_matrix = (
        np.outer(
            row_totals,
            col_totals,
        )
        / N
    )

    expected = (
        W * expected_matrix
    ).sum()

    if expected == 0:
        return 0.0

    return float(
        1
        - observed
        / expected
    )


# ============================================================
# FORECAST CLASSIFICATION
# ============================================================

def sat_index_array(values, thresholds):
    """
    Repository-equivalent forecast classification:

        score <= low       -> No Impact (0)
        score <= moderate  -> Low       (1)
        score <= high      -> Moderate  (2)
        score > high       -> High      (3)

    The thresholds supplied here are already rounded to 4 decimals,
    matching the current DDM Kappa workflow.
    """
    values = np.asarray(
        values,
        dtype=float,
    )

    return np.where(
        values <= thresholds["low"],
        0,
        np.where(
            values <= thresholds["moderate"],
            1,
            np.where(
                values <= thresholds["high"],
                2,
                3,
            ),
        ),
    ).astype(int)


def load_forecast(
    file_path,
    sheet_name,
):
    """
    Load one forecast lead and derive the SAME lead-specific operational
    composite severity boundaries used by the repository.

    Excel X  = agriculture forecast class
    Excel AG = household forecast class

    Thresholds are rounded to 4 decimals BEFORE classification, exactly
    as in the current repository Kappa-calibration code.
    """
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
    )

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    agri_class_col = (
        df.columns[
            column_index_from_string("X") - 1
        ]
    )

    house_class_col = (
        df.columns[
            column_index_from_string("AG") - 1
        ]
    )

    house_class = (
        df[house_class_col]
        .astype(str)
        .str.strip()
        .str.lower()
    )

    agri_class = (
        df[agri_class_col]
        .astype(str)
        .str.strip()
        .str.lower()
    )

    numeric_cols = [
        "Norm_Impact_House",
        "Norm_Impact_fAPAR",
        "Norm_Vul",
        "Norm_Wind Gust",
        "Norm_Rainfall",
        "Norm_Storm Surge",
    ]

    missing = [
        c
        for c in numeric_cols
        if c not in df.columns
    ]

    if missing:
        raise ValueError(
            f"Missing forecast columns in {file_path}: "
            + ", ".join(missing)
        )

    for col in numeric_cols:
        df[col] = pd.to_numeric(
            df[col],
            errors="coerce",
        )

    # --------------------------------------------------------
    # EXACT REPOSITORY FORECAST THRESHOLD DERIVATION
    # --------------------------------------------------------

    thresholds = {
        "house": {
            "low": round(
                df.loc[
                    house_class == "no impact",
                    "Norm_Impact_House",
                ].max(),
                4,
            ),
            "moderate": round(
                df.loc[
                    house_class == "low",
                    "Norm_Impact_House",
                ].max(),
                4,
            ),
            "high": round(
                df.loc[
                    house_class == "moderate",
                    "Norm_Impact_House",
                ].max(),
                4,
            ),
        },

        "agri": {
            "low": round(
                df.loc[
                    agri_class == "no impact",
                    "Norm_Impact_fAPAR",
                ].max(),
                4,
            ),
            "moderate": round(
                df.loc[
                    agri_class == "low",
                    "Norm_Impact_fAPAR",
                ].max(),
                4,
            ),
            "high": round(
                df.loc[
                    agri_class == "moderate",
                    "Norm_Impact_fAPAR",
                ].max(),
                4,
            ),
        },
    }

    for sector, t in thresholds.items():

        if any(
            pd.isna(v)
            for v in t.values()
        ):
            raise ValueError(
                f"Could not derive {sector} forecast "
                f"thresholds from {file_path}: {t}"
            )

        if not (
            t["low"]
            <= t["moderate"]
            <= t["high"]
        ):
            raise ValueError(
                f"Invalid {sector} forecast threshold "
                f"order in {file_path}: {t}"
            )

    # --------------------------------------------------------
    # USER-REQUESTED WEIGHTED HAZARD
    #
    # H*0.35 + J*0.35 + L*0.30
    # --------------------------------------------------------

    df["Weighted_Hazard"] = (
        0.35 * df["Norm_Wind Gust"]
        + 0.35 * df["Norm_Rainfall"]
        + 0.30 * df["Norm_Storm Surge"]
    )

    # Same admin cleaning / duplicate handling as repository.
    df = (
        df
        .dropna(
            subset=[
                "District",
                "Upazila",
            ]
        )
        .drop_duplicates(
            subset=[
                "District",
                "Upazila",
            ],
            keep="first",
        )
        .reset_index(
            drop=True,
        )
    )

    for col in [
        "District",
        "Upazila",
    ]:
        df[col] = (
            df[col]
            .apply(clean_admin)
        )

    return df, thresholds


# ============================================================
# LOAD DDM
# ============================================================

def load_ddm(
    ddm_file,
    house_sheet,
    agri_sheet,
):
    """
    Same DDM column extraction and cleaning as repository workflow.
    """

    # --------------------------------------------------------
    # HOUSE
    # --------------------------------------------------------

    raw_h = pd.read_excel(
        ddm_file,
        sheet_name=house_sheet,
        header=None,
    )

    house = (
        raw_h
        .iloc[
            2:,
            [0, 1, 5, 9],
        ]
        .copy()
        .reset_index(drop=True)
    )

    house.columns = [
        "District",
        "Upazila",
        "No_Total",
        "Amt_Total",
    ]

    house["District"] = (
        house["District"]
        .ffill()
    )

    house = (
        house
        .dropna(
            subset=[
                "District",
                "Upazila",
            ]
        )
    )

    for col in [
        "No_Total",
        "Amt_Total",
    ]:
        house[col] = (
            pd.to_numeric(
                house[col],
                errors="coerce",
            )
            .fillna(0)
        )

    for col in [
        "District",
        "Upazila",
    ]:
        house[col] = (
            house[col]
            .apply(clean_admin)
        )

    # --------------------------------------------------------
    # AGRICULTURE
    # --------------------------------------------------------

    raw_a = pd.read_excel(
        ddm_file,
        sheet_name=agri_sheet,
        header=None,
    )

    agri = (
        raw_a
        .iloc[
            2:,
            [0, 1, 6, 7],
        ]
        .copy()
        .reset_index(drop=True)
    )

    agri.columns = [
        "District",
        "Upazila",
        "Total_Loss_Land",
        "Total_Loss_Amt",
    ]

    agri["District"] = (
        agri["District"]
        .ffill()
    )

    agri = (
        agri
        .dropna(
            subset=[
                "District",
                "Upazila",
            ]
        )
    )

    for col in [
        "Total_Loss_Land",
        "Total_Loss_Amt",
    ]:
        agri[col] = (
            pd.to_numeric(
                agri[col],
                errors="coerce",
            )
            .fillna(0)
        )

    for col in [
        "District",
        "Upazila",
    ]:
        agri[col] = (
            agri[col]
            .apply(clean_admin)
        )

    return house, agri


# ============================================================
# DDM SEVERITY CLASSIFICATION
# ============================================================

def ddm_index_array(
    damage,
    low_cut,
    high_cut,
):
    """
    Same DDM severity classification as repository:

        damage <= 0          -> No Impact (0)
        damage <= low_cut    -> Low       (1)
        damage <= high_cut   -> Medium    (2)
        damage > high_cut    -> High      (3)
    """
    damage = np.asarray(
        damage,
        dtype=float,
    )

    return np.where(
        damage <= 0,
        0,
        np.where(
            damage <= low_cut,
            1,
            np.where(
                damage <= high_cut,
                2,
                3,
            ),
        ),
    ).astype(int)


# ============================================================
# EXACT REPOSITORY THREE-LEAD DDM CUT SEARCH
# ============================================================

def find_optimal_cuts_multilead(
    lead_arrays,
    candidate_damage,
):
    """
    For every valid pair of unique positive observed DDM values:

        1. classify DDM for 1dlt
        2. calculate QWK for 1dlt
        3. repeat for 2dlt and 3dlt
        4. average the three QWK values
        5. keep the pair with maximum mean QWK

    high_cut is never allowed to equal the maximum observed positive
    DDM value, preserving at least one High case.

    Tie handling also follows the repository: update only when
    mean_kappa > best_mean, so the first maximum encountered is kept.
    """

    candidate_damage = np.asarray(
        candidate_damage,
        dtype=float,
    )

    nz_vals = np.sort(
        np.unique(
            candidate_damage[
                candidate_damage > 0
            ]
        )
    )

    n_vals = len(nz_vals)

    if n_vals < 3:
        return (
            np.nan,
            np.nan,
            np.nan,
            {},
            pd.DataFrame(),
        )

    best_mean = -np.inf
    best_low = np.nan
    best_high = np.nan
    best_kappas = {}

    rows = []

    for i, low_cut in enumerate(
        nz_vals[:-2]
    ):

        for high_cut in (
            nz_vals[
                i + 1:-1
            ]
        ):

            kappas = {}

            for lead in LEADS:

                sat_idx = (
                    lead_arrays[lead]["sat"]
                )

                dmg = (
                    lead_arrays[lead]["dmg"]
                )

                ddm_idx = (
                    ddm_index_array(
                        dmg,
                        low_cut,
                        high_cut,
                    )
                )

                kappas[lead] = (
                    weighted_kappa(
                        sat_idx,
                        ddm_idx,
                    )
                )

            mean_kappa = float(
                np.nanmean([
                    kappas["1dlt"],
                    kappas["2dlt"],
                    kappas["3dlt"],
                ])
            )

            rows.append({
                "Low_Cut": float(low_cut),
                "High_Cut": float(high_cut),
                "Kappa_1dlt": kappas["1dlt"],
                "Kappa_2dlt": kappas["2dlt"],
                "Kappa_3dlt": kappas["3dlt"],
                "Mean_Kappa": mean_kappa,
            })

            if mean_kappa > best_mean:

                best_mean = mean_kappa
                best_low = float(low_cut)
                best_high = float(high_cut)
                best_kappas = kappas.copy()

    sweep = (
        pd.DataFrame(rows)
        .sort_values(
            "Mean_Kappa",
            ascending=False,
        )
        .reset_index(drop=True)
    )

    return (
        best_low,
        best_high,
        float(best_mean),
        best_kappas,
        sweep,
    )


# ============================================================
# AGREEMENT TABLE STATISTICS
# ============================================================

def agreement_stats(
    sat_idx,
    damage,
    low_cut,
    high_cut,
):
    """
    Kappa uses the same four-class arrays as repository.

    For the requested comparison table, agreement is partitioned into:

        Exact      : abs(forecast - observed) == 0
        Adjacent   : abs(forecast - observed) == 1
        Off-Cat.   : abs(forecast - observed) >= 2

    These are mutually exclusive and therefore sum to 100%.

    The repository's Summary field called "Within-1 %" is:
        Exact % + Adjacent %
    and is also returned for auditing.
    """

    sat_idx = np.asarray(
        sat_idx,
        dtype=int,
    )

    ddm_idx = (
        ddm_index_array(
            damage,
            low_cut,
            high_cut,
        )
    )

    diff = np.abs(
        sat_idx
        - ddm_idx
    )

    n = len(diff)

    if n == 0:
        return {
            "n": 0,
            "kappa": np.nan,
            "exact": np.nan,
            "adjacent": np.nan,
            "within1": np.nan,
            "off": np.nan,
        }

    exact = (
        np.mean(diff == 0)
        * 100
    )

    adjacent = (
        np.mean(diff == 1)
        * 100
    )

    off = (
        np.mean(diff >= 2)
        * 100
    )

    within1 = (
        np.mean(diff <= 1)
        * 100
    )

    return {
        "n": int(n),
        "kappa": weighted_kappa(
            sat_idx,
            ddm_idx,
        ),
        "exact": float(exact),
        "adjacent": float(adjacent),
        "within1": float(within1),
        "off": float(off),
    }


# ============================================================
# FORMAT HELPERS
# ============================================================

def format_cut_value(v):
    """
    Compact display similar to the supplied required-result table.
    """
    if pd.isna(v):
        return "NA"

    v = float(v)

    if abs(v - round(v)) < 1e-10:
        return f"{int(round(v)):,}"

    # retain sensible decimals without trailing zeros
    return (
        f"{v:,.2f}"
        .rstrip("0")
        .rstrip(".")
    )


def format_cut_pair(low_cut, high_cut):
    return (
        f"{format_cut_value(low_cut)} / "
        f"{format_cut_value(high_cut)}"
    )


def excel_safe_sheet_name(name):
    bad = '[]:*?/\\'
    for ch in bad:
        name = name.replace(ch, "_")
    return name[:31]


# ============================================================
# WRITE ONE REQUIRED-RESULT STYLE TABLE
# ============================================================

TABLE_HEADERS = [
    "Predictor",
    "Lead",
    "n",
    "DDM cut-points (Low / High)",
    "Kappa",
    "Exact %",
    "Adjacent %",
    "Off-Cat. %",
]


def write_required_table(
    ws,
    cyclone,
    variable_label,
    table_df,
):
    """
    Write a table with the same visual/content structure as the user's
    supplied required-result example.
    """

    last_col = (
        get_column_letter(
            len(TABLE_HEADERS)
        )
    )

    ws.merge_cells(
        f"A1:{last_col}1"
    )

    ws["A1"] = (
        f"IBF-Analysis - "
        f"Hazard/Vulnerability Ablation Justification"
    )

    ws["A1"].font = Font(
        name="Arial",
        size=12,
        bold=True,
        color=WHITE,
    )

    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = LEFT

    ws.merge_cells(
        f"A2:{last_col}2"
    )

    ws["A2"] = (
        f"Detail - {cyclone}, {variable_label}"
    )

    ws["A2"].font = Font(
        name="Arial",
        size=10,
        bold=True,
        color="000000",
    )

    ws["A2"].fill = fill(LBLUE)
    ws["A2"].alignment = LEFT

    for ci, h in enumerate(
        TABLE_HEADERS,
        1,
    ):
        c = ws.cell(
            row=4,
            column=ci,
            value=h,
        )
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = CENTER
        c.border = TB

    row_no = 5

    predictor_order = [
        "Composite",
        "Vulnerability alone",
        "Hazard-only (weighted)",
    ]

    lead_order = {
        "1dlt": 0,
        "2dlt": 1,
        "3dlt": 2,
    }

    data = table_df.copy()

    data["_pred_order"] = (
        data["Predictor"]
        .map({
            p: i
            for i, p in enumerate(
                predictor_order
            )
        })
    )

    data["_lead_order"] = (
        data["Lead"]
        .map(lead_order)
    )

    data = (
        data
        .sort_values([
            "_pred_order",
            "_lead_order",
        ])
        .reset_index(drop=True)
    )

    for _, row in data.iterrows():

        values = [
            row["Predictor"],
            row["Lead"],
            int(row["n"]),
            row["DDM cut-points (Low / High)"],
            float(row["Kappa"]),
            float(row["Exact %"]),
            float(row["Adjacent %"]),
            float(row["Off-Cat. %"]),
        ]

        for ci, value in enumerate(
            values,
            1,
        ):

            c = ws.cell(
                row=row_no,
                column=ci,
                value=value,
            )

            c.font = body_font()
            c.border = TB

            c.alignment = (
                LEFT
                if ci == 1
                else CENTER
            )

            c.fill = fill(
                LGREY
                if row_no % 2 == 0
                else WHITE
            )

            if ci == 5:
                c.number_format = "0.000"
                c.fill = fill(
                    kappa_fill(
                        float(value)
                    )
                )

            if ci in [6, 7, 8]:
                c.number_format = "0.0"

        row_no += 1

    widths = {
        "A": 27,
        "B": 10,
        "C": 8,
        "D": 26,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 13,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = (
        f"A4:{last_col}{row_no - 1}"
    )


# ============================================================
# PROCESS ONE CYCLONE
# ============================================================

def run_cyclone(
    project_root,
    cyclone_key,
):
    config_file = (
        project_root
        / "configs"
        / f"{cyclone_key}.yaml"
    )

    if not config_file.is_file():
        raise FileNotFoundError(
            f"Config not found: "
            f"{config_file}"
        )

    with open(
        config_file,
        "r",
        encoding="utf-8",
    ) as f:
        config = yaml.safe_load(f)

    cyclone = (
        config["cyclone"]["name"]
    )

    forecast_sheet = (
        config["sheets"]["forecast"]
    )

    house_sheet = (
        config["sheets"]["house"]
    )

    agri_sheet = (
        config["sheets"]["agriculture"]
    )

    ddm_file = (
        project_root
        / config["data"]["observed"]
    )

    forecast_files = {
        lead:
            project_root
            / config["data"]["forecasts"][lead]
        for lead in LEADS
    }

    # --------------------------------------------------------
    # LOAD ALL THREE FORECAST LEADS
    # --------------------------------------------------------

    forecasts = {}
    thresholds_by_lead = {}

    print()
    print("=" * 76)
    print(
        f"{cyclone} - "
        "HAZARD/VULNERABILITY ABLATION KAPPA"
    )
    print("=" * 76)

    for lead in LEADS:

        (
            forecasts[lead],
            thresholds_by_lead[lead],
        ) = load_forecast(
            forecast_files[lead],
            forecast_sheet,
        )

        print()
        print(lead)
        print(
            "  House thresholds:",
            thresholds_by_lead[lead]["house"],
        )
        print(
            "  Agriculture thresholds:",
            thresholds_by_lead[lead]["agri"],
        )

    # --------------------------------------------------------
    # LOAD DDM
    # --------------------------------------------------------

    (
        ddm_house,
        ddm_agri,
    ) = load_ddm(
        ddm_file,
        house_sheet,
        agri_sheet,
    )

    # --------------------------------------------------------
    # MERGE EXACTLY AS REPOSITORY: District + Upazila
    # --------------------------------------------------------

    merged = {}

    print()
    print("MERGED RECORDS")
    print("-" * 76)

    for lead in LEADS:

        house = pd.merge(
            forecasts[lead],
            ddm_house,
            on=[
                "District",
                "Upazila",
            ],
            how="inner",
        )

        agri = pd.merge(
            forecasts[lead],
            ddm_agri,
            on=[
                "District",
                "Upazila",
            ],
            how="inner",
        )

        merged[lead] = {
            "house": house,
            "agri": agri,
        }

        print(
            f"{lead}: "
            f"House={len(house)}   "
            f"Agriculture={len(agri)}"
        )

    # --------------------------------------------------------
    # SAME FOUR REPOSITORY DAMAGE COMPARISONS
    # --------------------------------------------------------

    comparisons = [
        {
            "key": "House_NoTotal",
            "variable":
                "Housing - No. Damaged Households",
            "short_sheet":
                "Housing_NoTotal",
            "sector": "house",
            "damage_col": "No_Total",
            "ddm_source": ddm_house,
            "composite_col":
                "Norm_Impact_House",
        },

        {
            "key": "House_AmtTotal",
            "variable":
                "Housing - Repair Amount (BDT)",
            "short_sheet":
                "Housing_Amt",
            "sector": "house",
            "damage_col": "Amt_Total",
            "ddm_source": ddm_house,
            "composite_col":
                "Norm_Impact_House",
        },

        {
            "key": "Agri_LossLand",
            "variable":
                "Agriculture - Land Loss (ha)",
            "short_sheet":
                "Agri_Land",
            "sector": "agri",
            "damage_col":
                "Total_Loss_Land",
            "ddm_source": ddm_agri,
            "composite_col":
                "Norm_Impact_fAPAR",
        },

        {
            "key": "Agri_LossAmt",
            "variable":
                "Agriculture - Loss Amount (BDT)",
            "short_sheet":
                "Agri_Amt",
            "sector": "agri",
            "damage_col":
                "Total_Loss_Amt",
            "ddm_source": ddm_agri,
            "composite_col":
                "Norm_Impact_fAPAR",
        },
    ]

    rows = []
    sweep_tables = {}

    # --------------------------------------------------------
    # RUN COMPOSITE + TWO ABLATIONS
    # --------------------------------------------------------

    for comp in comparisons:

        print()
        print("-" * 76)
        print(comp["variable"])

        candidate_damage = (
            pd.to_numeric(
                comp["ddm_source"][
                    comp["damage_col"]
                ],
                errors="coerce",
            )
            .fillna(0)
            .values
        )

        for (
            predictor_name,
            predictor_col,
        ) in PREDICTORS:

            lead_arrays = {}

            for lead in LEADS:

                df = (
                    merged[lead][
                        comp["sector"]
                    ]
                    .copy()
                )

                thresholds = (
                    thresholds_by_lead[
                        lead
                    ][
                        comp["sector"]
                    ]
                )

                if predictor_name == "Composite":

                    score = (
                        pd.to_numeric(
                            df[
                                comp[
                                    "composite_col"
                                ]
                            ],
                            errors="coerce",
                        )
                    )

                else:

                    score = (
                        pd.to_numeric(
                            df[
                                predictor_col
                            ],
                            errors="coerce",
                        )
                    )

                damage = (
                    pd.to_numeric(
                        df[
                            comp[
                                "damage_col"
                            ]
                        ],
                        errors="coerce",
                    )
                    .fillna(0)
                )

                valid = (
                    score.notna()
                )

                score_valid = (
                    score[valid]
                )

                damage_valid = (
                    damage[valid]
                )

                sat_idx = (
                    sat_index_array(
                        score_valid.values,
                        thresholds,
                    )
                )

                lead_arrays[lead] = {
                    "sat": sat_idx,
                    "dmg":
                        damage_valid.values,
                }

            (
                fixed_low,
                fixed_high,
                mean_kappa,
                lead_kappas,
                sweep,
            ) = (
                find_optimal_cuts_multilead(
                    lead_arrays,
                    candidate_damage,
                )
            )

            if (
                pd.isna(fixed_low)
                or
                pd.isna(fixed_high)
            ):
                raise ValueError(
                    "Could not derive fixed DDM "
                    f"cuts for {cyclone}, "
                    f"{comp['variable']}, "
                    f"{predictor_name}"
                )

            sweep_tables[
                (
                    comp["key"],
                    predictor_name,
                )
            ] = sweep

            print(
                f"  {predictor_name:<24} "
                f"cuts = "
                f"{format_cut_pair(fixed_low, fixed_high):<22} "
                f"mean Kappa = "
                f"{mean_kappa:+.4f}"
            )

            for lead in LEADS:

                stats = (
                    agreement_stats(
                        lead_arrays[
                            lead
                        ]["sat"],
                        lead_arrays[
                            lead
                        ]["dmg"],
                        fixed_low,
                        fixed_high,
                    )
                )

                # Kappa must equal the value obtained during the
                # common-cut search. Verify numerically.
                if not np.isclose(
                    stats["kappa"],
                    lead_kappas[lead],
                    rtol=1e-12,
                    atol=1e-12,
                    equal_nan=True,
                ):
                    raise RuntimeError(
                        "Internal Kappa consistency "
                        f"check failed for "
                        f"{cyclone} / "
                        f"{comp['variable']} / "
                        f"{predictor_name} / "
                        f"{lead}"
                    )

                rows.append({
                    "Cyclone": cyclone,
                    "Variable":
                        comp["variable"],
                    "Variable_Key":
                        comp["key"],
                    "Predictor":
                        predictor_name,
                    "Lead": lead,
                    "n": stats["n"],
                    "DDM Low Cut":
                        fixed_low,
                    "DDM High Cut":
                        fixed_high,
                    "DDM cut-points (Low / High)":
                        format_cut_pair(
                            fixed_low,
                            fixed_high,
                        ),
                    "Kappa":
                        lead_kappas[lead],
                    "Mean Kappa":
                        mean_kappa,
                    "Exact %":
                        stats["exact"],
                    "Adjacent %":
                        stats["adjacent"],
                    "Within-1 %":
                        stats["within1"],
                    "Off-Cat. %":
                        stats["off"],
                })

    result = pd.DataFrame(rows)

    # --------------------------------------------------------
    # OUTPUT DIRECTORY
    # --------------------------------------------------------

    out_dir = (
        project_root
        / config["output"]["directory"]
        / "Ablation_Kappa_Comparison"
    )

    out_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    excel_path = (
        out_dir
        / f"{cyclone}_Ablation_Kappa_Comparison.xlsx"
    )

    # --------------------------------------------------------
    # EXCEL WORKBOOK
    # --------------------------------------------------------

    wb = Workbook()
    wb.remove(wb.active)

    # Main combined table
    ws_all = wb.create_sheet(
        "All_Comparisons"
    )

    all_headers = [
        "Cyclone",
        "Variable",
        "Predictor",
        "Lead",
        "n",
        "DDM cut-points (Low / High)",
        "Kappa",
        "Exact %",
        "Adjacent %",
        "Off-Cat. %",
    ]

    for ci, h in enumerate(
        all_headers,
        1,
    ):
        c = ws_all.cell(
            row=1,
            column=ci,
            value=h,
        )
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = CENTER
        c.border = TB

    for ri, (
        _,
        row,
    ) in enumerate(
        result.iterrows(),
        2,
    ):

        vals = [
            row["Cyclone"],
            row["Variable"],
            row["Predictor"],
            row["Lead"],
            int(row["n"]),
            row[
                "DDM cut-points (Low / High)"
            ],
            float(row["Kappa"]),
            float(row["Exact %"]),
            float(row["Adjacent %"]),
            float(row["Off-Cat. %"]),
        ]

        for ci, v in enumerate(
            vals,
            1,
        ):
            c = ws_all.cell(
                row=ri,
                column=ci,
                value=v,
            )
            c.border = TB
            c.font = body_font()
            c.alignment = (
                LEFT
                if ci in [1, 2, 3]
                else CENTER
            )
            c.fill = fill(
                LGREY
                if ri % 2 == 0
                else WHITE
            )

            if ci == 7:
                c.number_format = "0.000"
                c.fill = fill(
                    kappa_fill(
                        float(v)
                    )
                )

            if ci in [8, 9, 10]:
                c.number_format = "0.0"

    widths = [
        14,
        34,
        26,
        10,
        8,
        26,
        12,
        12,
        14,
        13,
    ]

    for ci, width in enumerate(
        widths,
        1,
    ):
        ws_all.column_dimensions[
            get_column_letter(ci)
        ].width = width

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = (
        f"A1:J{len(result) + 1}"
    )

    # One exact required-result-style sheet per damage variable
    for comp in comparisons:

        sub = (
            result[
                result["Variable_Key"]
                == comp["key"]
            ]
            .copy()
        )

        ws = wb.create_sheet(
            excel_safe_sheet_name(
                comp["short_sheet"]
            )
        )

        write_required_table(
            ws,
            cyclone,
            comp["variable"],
            sub,
        )

    # Audit sheet to prove the reporting transformation:
    # Within-1 = Exact + Adjacent
    ws_audit = wb.create_sheet(
        "Audit"
    )

    audit_headers = [
        "Variable",
        "Predictor",
        "Lead",
        "Kappa",
        "Exact %",
        "Adjacent %",
        "Within-1 %",
        "Exact + Adjacent",
        "Off-Cat. %",
        "Sum Exact+Adjacent+Off",
        "Check",
    ]

    for ci, h in enumerate(
        audit_headers,
        1,
    ):
        c = ws_audit.cell(
            1,
            ci,
            h,
        )
        c.font = header_font()
        c.fill = fill(NAVY)
        c.alignment = CENTER
        c.border = TB

    for ri, (
        _,
        row,
    ) in enumerate(
        result.iterrows(),
        2,
    ):

        exact_adj = (
            row["Exact %"]
            + row["Adjacent %"]
        )

        total = (
            exact_adj
            + row["Off-Cat. %"]
        )

        ok = (
            np.isclose(
                exact_adj,
                row["Within-1 %"],
                atol=1e-9,
            )
            and
            np.isclose(
                total,
                100.0,
                atol=1e-9,
            )
        )

        vals = [
            row["Variable"],
            row["Predictor"],
            row["Lead"],
            row["Kappa"],
            row["Exact %"],
            row["Adjacent %"],
            row["Within-1 %"],
            exact_adj,
            row["Off-Cat. %"],
            total,
            "OK" if ok else "CHECK",
        ]

        for ci, v in enumerate(
            vals,
            1,
        ):
            c = ws_audit.cell(
                ri,
                ci,
                v,
            )
            c.border = TB
            c.font = body_font()
            c.alignment = (
                LEFT
                if ci in [1, 2]
                else CENTER
            )

            if ci == 4:
                c.number_format = "0.0000"

            if ci in [
                5, 6, 7, 8, 9, 10
            ]:
                c.number_format = "0.0000"

    for ci in range(
        1,
        len(audit_headers) + 1,
    ):
        ws_audit.column_dimensions[
            get_column_letter(ci)
        ].width = (
            34
            if ci == 1
            else
            24
            if ci == 2
            else
            16
        )

    ws_audit.freeze_panes = "A2"

    # Method sheet
    ws_method = wb.create_sheet(
        "Method"
    )

    method_lines = [
        "Ablation comparison method",
        "",
        "Composite:",
        "Uses Norm_Impact_House for housing and Norm_Impact_fAPAR for agriculture.",
        "",
        "Vulnerability alone:",
        "Uses Norm_Vul, Excel column F.",
        "",
        "Hazard-only weighted:",
        "Weighted_Hazard = 0.35*Norm_Wind Gust + 0.35*Norm_Rainfall + 0.30*Norm_Storm Surge",
        "Equivalent Excel columns: 0.35*H + 0.35*J + 0.30*L.",
        "",
        "Forecast severity boundaries:",
        "Derived independently for each lead exactly as in the repository.",
        "House: from Norm_Impact_House using Excel AG class labels.",
        "Agriculture: from Norm_Impact_fAPAR using Excel X class labels.",
        "Thresholds are rounded to 4 decimals before classification.",
        "The same lead-specific sector boundary set is applied to Composite, Vulnerability-only and Hazard-only.",
        "",
        "DDM calibration:",
        "Zero = No Impact.",
        "Positive DDM values are exhaustively searched for Low/High cut pairs.",
        "For each candidate pair, QWK is calculated separately for 1dlt, 2dlt and 3dlt.",
        "The mean of the three Kappas is maximized.",
        "The selected Low/High pair is then fixed for all three leads for that predictor and damage variable.",
        "The maximum positive observed DDM value is excluded as a High cut so that the High class remains populated.",
        "",
        "Agreement table:",
        "Exact = difference of 0 classes.",
        "Adjacent = difference of exactly 1 class.",
        "Off-Cat. = difference of 2 or more classes.",
        "Exact + Adjacent + Off-Cat. = 100%.",
        "Repository Within-1 % = Exact % + Adjacent %.",
    ]

    for r, text in enumerate(
        method_lines,
        1,
    ):
        c = ws_method.cell(
            r,
            1,
            text,
        )
        c.font = body_font(
            bold=(
                text.endswith(":")
                or
                text
                == "Ablation comparison method"
            )
        )
        c.alignment = LEFT

    ws_method.column_dimensions[
        "A"
    ].width = 120

    # Save workbook
    wb.save(
        excel_path
    )

    print()
    print("=" * 76)
    print("ABLATION KAPPA COMPARISON COMPLETED")
    print("=" * 76)
    print(f"Excel: {excel_path}")

    return result


# ============================================================
# MAIN
# ============================================================

def main():

    parser = argparse.ArgumentParser(
        description=(
            "Compare composite IBF severity skill "
            "with vulnerability-only and weighted "
            "hazard-only predictors using the "
            "repository's cyclone-wide fixed DDM "
            "QWK calibration."
        )
    )

    parser.add_argument(
        "--cyclone",
        choices=[
            "remal",
            "sitrang",
            "midhili",
            "all",
        ],
        default="all",
        help=(
            "Cyclone to process. "
            "Default: all (Remal, Midhili and Sitrang)"
        ),
    )

    args = parser.parse_args()

    project_root = (
        Path(__file__)
        .resolve()
        .parents[1]
    )

    if args.cyclone == "all":

        for cyclone_key in [
            "remal",
            "midhili",
            "sitrang",
        ]:
            run_cyclone(
                project_root,
                cyclone_key,
            )

    else:

        run_cyclone(
            project_root,
            args.cyclone,
        )


if __name__ == "__main__":
    main()
